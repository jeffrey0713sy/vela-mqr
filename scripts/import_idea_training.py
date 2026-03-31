"""
Import idea_training.xlsx -> data/idea_training/idea_training.csv + manifest.json

Requires: pip install openpyxl (see requirements-idea.txt)
"""

from __future__ import annotations

import argparse
import csv
import json
import os
from datetime import datetime
from pathlib import Path


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def main() -> None:
    parser = argparse.ArgumentParser(description="Export idea_training sheet to CSV.")
    parser.add_argument(
        "--input",
        type=str,
        default="",
        help="Path to idea_training.xlsx (or set IDEA_TRAINING_XLSX)",
    )
    parser.add_argument(
        "--output",
        type=str,
        default="data/idea_training/idea_training.csv",
    )
    parser.add_argument(
        "--sheet",
        type=str,
        default="idea_training",
    )
    args = parser.parse_args()

    src = (args.input or os.environ.get("IDEA_TRAINING_XLSX", "")).strip()
    if not src:
        raise SystemExit("Pass --input path/to/idea_training.xlsx or set IDEA_TRAINING_XLSX")

    in_path = Path(src).expanduser().resolve()
    if not in_path.is_file():
        raise FileNotFoundError(in_path)

    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit("Install openpyxl: pip install -r requirements-idea.txt") from e

    root = Path(__file__).resolve().parents[1]
    out_path = (root / args.output).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(in_path, read_only=True, data_only=True)
    if args.sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet {args.sheet!r} not in {wb.sheetnames}")
    ws = wb[args.sheet]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    if not header:
        raise SystemExit("Empty sheet")
    header = [
        str(h).strip().lower().replace(" ", "_") if h is not None else f"col_{i}"
        for i, h in enumerate(header)
    ]

    n = 0
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for row in rows_iter:
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            padded = list(row) + [None] * (len(header) - len(row))
            padded = padded[: len(header)]
            w.writerow([_cell_str(c) for c in padded])
            n += 1

    wb.close()

    manifest = {
        "source_xlsx": str(in_path),
        "sheet": args.sheet,
        "output_csv": str(out_path.relative_to(root)),
        "n_rows": n,
        "exported_at": datetime.now().isoformat(timespec="seconds"),
    }
    man_path = out_path.parent / "manifest.json"
    man_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    print(f"Wrote {n} rows to {out_path}")
    print(f"Manifest: {man_path}")


if __name__ == "__main__":
    main()
